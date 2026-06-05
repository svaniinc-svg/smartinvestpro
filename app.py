# app.py
import streamlit as st
import pandas as pd
import numpy as np
import numpy_financial as npf
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import matplotlib.pyplot as plt
from datetime import datetime
import re
from pathlib import Path
import uuid


st.set_page_config(page_title="Real Estate ROI Calculator", layout="wide")
st.title("🏡 Smart Rental ROI Calculator + Investor Lead Funnel")

st.caption("No signup required. Run the numbers first. Join the investor deals list later if you want Charlotte-area opportunities and deal review follow-up.")

# -------------------------
# Financial helpers
# -------------------------
def pmt(monthly_rate, nper, pv):
    if nper <= 0:
        return 0.0
    if abs(monthly_rate) < 1e-12:
        return pv / nper
    return monthly_rate * pv / (1 - (1 + monthly_rate) ** (-nper))

def _npv_scalar(rate, cashflows):
    t = np.arange(len(cashflows), dtype=float)
    return float(np.sum(np.array(cashflows, dtype=float) / np.power(1.0 + rate, t)))

def _irr_stable(cashflows, guess=0.01):
    cf = np.array(cashflows, dtype=float)
    if not (np.any(cf > 0) and np.any(cf < 0)):
        return None
    # Try polynomial roots first
    try:
        coeffs = cf[::-1]
        roots = np.roots(coeffs)
        candidates = []
        for r in roots:
            if abs(r.imag) < 1e-10 and r.real != 0:
                x = float(r.real)
                if x > 0:
                    rate = (1.0 / x) - 1.0
                    if rate > -0.999999:
                        candidates.append(rate)
        if candidates:
            return min(candidates, key=lambda z: abs(z - guess))
    except Exception:
        pass
    # Bracket + bisection on per-period IRR
    low, high = -0.999999, 10.0
    f_low, f_high = _npv_scalar(low, cf), _npv_scalar(high, cf)
    tries = 0
    while f_low * f_high > 0 and tries < 60:
        high *= 2.0
        f_high = _npv_scalar(high, cf)
        tries += 1
    if f_low * f_high > 0:
        return None
    for _ in range(200):
        mid = (low + high) / 2.0
        f_mid = _npv_scalar(mid, cf)
        if abs(f_mid) < 1e-10:
            return mid
        if f_low * f_mid < 0:
            high, f_high = mid, f_mid
        else:
            low, f_low = mid, f_mid
    return (low + high) / 2.0

def _annualize_from_monthly(r):
    return (1 + r) ** 12 - 1 if r is not None and not np.isnan(r) else None

# -------------------------
# Styling helpers
# -------------------------
def _style_red_neg(val):
    try:
        if isinstance(val, (int, float, np.number)) and val < 0:
            return "color:red;"
    except Exception:
        pass
    return ""

def _fmt_metric(metric, val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "N/A"
    percent_keys = ["Cap Rate", "Cash on Cash", "IRR", "ROI", "Effective Rate", "Occupancy"]
    currency_keys = ["Price", "Payment", "Subsidy", "NOI", "Debt Service", "Cash Flow",
                     "Loan", "Proceeds", "Equity", "Profit", "Investment", "Closing Costs", "Commission", "Balance"]
    try:
        if any(k in metric for k in percent_keys):
            return f"{val*100:.2f}%"
        if any(k in metric for k in currency_keys):
            return f"${val:,.2f}"
        if metric.startswith("DCR"):
            return f"{val:.2f}"
        if isinstance(val, (int, float, np.number)):
            return f"{val:,.2f}"
    except Exception:
        pass
    return val

# -------------------------
# Buydown helpers
# -------------------------
def _buydown_preview(loan_amount, annual_rate, nper):
    """Return Yr1/Yr2 effective rates/payments and subsidy totals."""
    if nper <= 0 or loan_amount <= 0:
        return (annual_rate, annual_rate, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0)
    note_pmt = pmt(annual_rate/12.0, nper, loan_amount)
    eff1 = max(annual_rate - 0.02, 0.0)
    eff2 = max(annual_rate - 0.01, 0.0)
    p1 = pmt(eff1/12.0, nper, loan_amount)
    p2 = pmt(eff2/12.0, nper, loan_amount)
    m1 = min(12, nper)
    m2 = min(max(nper - 12, 0), 12)
    sub1 = (note_pmt - p1) * m1
    sub2 = (note_pmt - p2) * m2
    return eff1, eff2, p1, p2, note_pmt, sub1, sub2, (sub1 + sub2)

def _buydown_oop_series(loan_amount, annual_rate, nper, note_payment, use_buydown_2_1):
    """Borrower out-of-pocket monthly payments (amortization still at note payment)."""
    if not use_buydown_2_1:
        return np.full(nper, note_payment, dtype=float)
    eff1 = max(annual_rate - 0.02, 0.0)
    eff2 = max(annual_rate - 0.01, 0.0)
    p1 = pmt(eff1 / 12.0, nper, loan_amount)
    p2 = pmt(eff2 / 12.0, nper, loan_amount)
    arr = []
    for m in range(1, nper + 1):
        if m <= 12:
            arr.append(p1)
        elif m <= 24:
            arr.append(p2)
        else:
            arr.append(note_payment)
    return np.array(arr, dtype=float)

# -------------------------
# Amortization + Metrics
# -------------------------
def amortization_schedule(loan_amount, annual_rate, years, down_payment=0, override_payment=None, use_buydown_2_1=False):
    """
    Returns a schedule where:
    - Note Payment drives Interest/Principal/Balance (note rate).
    - Borrower Payment reflects 2/1 buydown out-of-pocket.
    - Subsidy = Note Payment - Borrower Payment.
    """
    monthly_rate = annual_rate / 12.0
    nper = int(round(years * 12))
    cols = ["Month", "Note Payment", "Borrower Payment", "Subsidy", "Principal", "Interest", "Balance", "Equity"]
    if nper <= 0:
        return pd.DataFrame(columns=cols)

    note_payment = override_payment if (override_payment and override_payment > 0) else pmt(monthly_rate, nper, loan_amount)
    oop = _buydown_oop_series(loan_amount, annual_rate, nper, note_payment, use_buydown_2_1)

    balance = float(loan_amount)
    rows = []
    for m in range(1, nper + 1):
        interest = balance * monthly_rate
        principal = note_payment - interest
        if m == nper:
            principal = balance
            true_note_payment = principal + interest
            balance = 0.0
        else:
            balance -= principal
            true_note_payment = note_payment
            if balance < 1e-8:
                balance = 0.0

        borrower_pay = float(oop[m - 1])
        subsidy = true_note_payment - borrower_pay
        equity = down_payment + (loan_amount - balance)

        rows.append({
            "Month": m,
            "Note Payment": round(true_note_payment, 2),
            "Borrower Payment": round(borrower_pay, 2),
            "Subsidy": round(subsidy, 2),
            "Principal": round(principal, 2),
            "Interest": round(interest, 2),
            "Balance": round(balance, 2),
            "Equity": round(equity, 2),
        })
    return pd.DataFrame(rows)

def build_yearly_cashflows(hold_years, nper, effective_rent_base, monthly_expenses, monthly_taxes, monthly_insurance,
                           monthly_hoa, mgmt_pct, rent_appreciation_pct, oop_series):
    """Returns a DataFrame of yearly cashflows (1..hold_years) honoring rent growth and partial first/last years."""
    rows = []
    cum_cf = 0.0
    for year in range(1, hold_years + 1):
        # Rent growth compounded annually
        rent_multiplier = (1 + rent_appreciation_pct / 100.0) ** (year - 1)
        eff_rent_y = effective_rent_base * rent_multiplier  # monthly effective rent (after vacancy)
        mgmt_fee_y = eff_rent_y * (mgmt_pct / 100.0)
        tot_mo_exp_y = monthly_expenses + monthly_taxes + monthly_insurance + monthly_hoa + mgmt_fee_y

        start = (year - 1) * 12
        end = min(year * 12, nper)
        months = max(0, end - start)
        if months == 0:
            break

        noi_year = (eff_rent_y - tot_mo_exp_y) * 12.0
        noi_prorated = noi_year * (months / 12.0)
        debt_service = float(oop_series[start:end].sum())
        cash_flow = noi_prorated - debt_service
        cum_cf += cash_flow

        rows.append({
            "Year": year,
            "Months Counted": months,
            "Gross Rent (after vacancy) /mo": round(eff_rent_y, 2),
            "Mgmt Fee /mo": round(mgmt_fee_y, 2),
            "Other OpEx /mo (excl taxes/ins/HOA)": round(monthly_expenses, 2),
            "Taxes /mo": round(monthly_taxes, 2),
            "Insurance /mo": round(monthly_insurance, 2),
            "HOA /mo": round(monthly_hoa, 2),
            "Total OpEx /mo": round(tot_mo_exp_y, 2),
            "NOI (annual, prorated)": round(noi_prorated, 2),
            "Debt Service (annual, prorated)": round(debt_service, 2),
            "Cash Flow (annual)": round(cash_flow, 2),
            "Cumulative Cash Flow": round(cum_cf, 2),
        })
    return pd.DataFrame(rows)

def calculate_metrics(purchase_price, down_payment, annual_rate, years,
                      monthly_rent, monthly_expenses, monthly_taxes, monthly_insurance, monthly_hoa,
                      vacancy_pct=0.0, mgmt_pct=0.0, hold_years=5,
                      rent_appreciation_pct=0.0, appreciation_pct=0.0,
                      buy_closing_costs=0.0, sale_closing_costs=0.0, sale_commission_pct=0.0,
                      override_payment=None, use_buydown_2_1=False, buydown_paid_by="Buyer",
                      sale_price_override=None, buy_cc_paid_by="Buyer"):
    loan_amount = max(0.0, purchase_price - down_payment)
    amort = amortization_schedule(
        loan_amount, annual_rate, years,
        down_payment=down_payment, override_payment=override_payment, use_buydown_2_1=use_buydown_2_1
    )
    nper = len(amort)
    monthly_note_payment = amort["Note Payment"].iloc[0] if nper > 0 else 0.0

    # Base monthly figures (Yr1)
    effective_rent = monthly_rent * (1 - vacancy_pct / 100.0)  # after vacancy
    mgmt_fee = effective_rent * (mgmt_pct / 100.0)
    total_monthly_expenses = monthly_expenses + monthly_taxes + monthly_insurance + monthly_hoa + mgmt_fee
    noi = (effective_rent - total_monthly_expenses) * 12.0  # Yr1 annual NOI

    # Out-of-pocket debt service series
    oop_series = amort["Borrower Payment"].to_numpy(dtype=float) if nper > 0 else np.array([])

    # Buydown preview/subsidy
    eff_rate_y1, eff_rate_y2, borrower_pmt_y1, borrower_pmt_y2, note_pmt_ref, subsidy_y1, subsidy_y2, subsidy_total = \
        _buydown_preview(loan_amount, annual_rate, nper) if use_buydown_2_1 else \
        (annual_rate, annual_rate, monthly_note_payment, monthly_note_payment, monthly_note_payment, 0.0, 0.0, 0.0)

    # Initial investment: down + buy CC (if Buyer) + buydown (if Buyer)
    initial_investment = down_payment
    if buy_cc_paid_by == "Buyer":
        initial_investment += buy_closing_costs
    if use_buydown_2_1 and buydown_paid_by == "Buyer":
        initial_investment += subsidy_total

    # Yr1 snapshots
    months_y1 = min(12, nper)
    months_y2 = min(max(nper - 12, 0), 12)
    annual_debt_y1 = float(oop_series[:months_y1].sum()) if nper > 0 else 0.0
    annual_cash_flow_y1 = noi - annual_debt_y1
    annual_debt_y2 = float(oop_series[12:12 + months_y2].sum()) if nper > 12 else None
    annual_cash_flow_y2 = (noi - annual_debt_y2) if nper > 12 else None

    # DCR (Yr1)
    dcr_y1 = (noi / annual_debt_y1) if annual_debt_y1 > 0 else None

    # Break-Even Occupancy % (Yr1) — fraction (e.g., 0.85 = 85%)
    gpi_annual = monthly_rent * 12.0
    mgmt_frac = (mgmt_pct / 100.0)
    fixed_opex_annual = (monthly_expenses + monthly_taxes + monthly_insurance + monthly_hoa) * 12.0
    denom = gpi_annual * (1.0 - mgmt_frac) if (1.0 - mgmt_frac) != 0 else np.nan
    break_even_occ = ((fixed_opex_annual + annual_debt_y1) / denom) if denom and not np.isnan(denom) else None

    # Display metrics
    cap_rate = (noi / purchase_price) if purchase_price else None
    cash_on_cash = (annual_cash_flow_y1 / initial_investment) if initial_investment > 0 else None

    # IRR across full term with sale
    term_years = int(round(years))
    annual_irr_term = None
    if term_years > 0 and nper > 0:
        cash_flows_term = [-initial_investment]
        for year in range(1, term_years + 1):
            rent_multiplier = (1 + rent_appreciation_pct / 100.0) ** (year - 1)
            eff_rent_y = effective_rent * rent_multiplier
            mgmt_fee_y = eff_rent_y * (mgmt_pct / 100.0)
            tot_mo_exp_y = monthly_expenses + monthly_taxes + monthly_insurance + monthly_hoa + mgmt_fee_y
            noi_y = (eff_rent_y - tot_mo_exp_y) * 12.0
            start = (year - 1) * 12
            end = min(year * 12, nper)
            months = end - start
            if months <= 0:
                continue
            annual_debt_y = float(oop_series[start:end].sum())
            cf_y = (noi_y * (months / 12.0)) - annual_debt_y
            cash_flows_term += [cf_y / months] * months

        sale_price_term = purchase_price * ((1 + appreciation_pct / 100.0) ** term_years)
        loan_bal_term = float(amort.iloc[-1]["Balance"]) if nper > 0 else loan_amount
        sale_commission_term = sale_price_term * (sale_commission_pct / 100.0)
        net_sale_term = sale_price_term - loan_bal_term - sale_commission_term - sale_closing_costs
        cash_flows_term.append(net_sale_term)

        try:
            per_period_irr = npf.irr(cash_flows_term)
            if per_period_irr is None or np.isnan(per_period_irr):
                per_period_irr = _irr_stable(np.array(cash_flows_term))
        except Exception:
            per_period_irr = _irr_stable(np.array(cash_flows_term))
        annual_irr_term = _annualize_from_monthly(per_period_irr)

    # Hold-period prediction & IRR
    hold_months = int(min(hold_years, years) * 12)
    if hold_months == 0 or len(amort) == 0:
        equity_at_hold = down_payment
        loan_balance_at_hold = loan_amount
    elif len(amort) >= hold_months:
        equity_at_hold = float(amort.loc[hold_months - 1, "Equity"])
        loan_balance_at_hold = float(amort.loc[hold_months - 1, "Balance"])
    else:
        equity_at_hold = down_payment
        loan_balance_at_hold = loan_amount

    yearly_cf_df = build_yearly_cashflows(
        hold_years=hold_years,
        nper=nper,
        effective_rent_base=effective_rent,
        monthly_expenses=monthly_expenses,
        monthly_taxes=monthly_taxes,
        monthly_insurance=monthly_insurance,
        monthly_hoa=monthly_hoa,
        mgmt_pct=mgmt_pct,
        rent_appreciation_pct=rent_appreciation_pct,
        oop_series=oop_series
    )

    cash_flows_hold = [-initial_investment]
    total_cash_flow_hold = 0.0
    for _, r in yearly_cf_df.iterrows():
        months = int(r["Months Counted"])
        if months > 0:
            cf = float(r["Cash Flow (annual)"])
            total_cash_flow_hold += cf
            cash_flows_hold += [cf / months] * months

    sale_price = float(sale_price_override) if (sale_price_override and sale_price_override > 0) else \
                 purchase_price * ((1 + appreciation_pct / 100.0) ** min(hold_years, years))
    sale_commission = sale_price * (sale_commission_pct / 100.0)
    net_sale_proceeds = sale_price - loan_balance_at_hold - sale_commission - sale_closing_costs
    cash_flows_hold.append(net_sale_proceeds)

    total_profit = net_sale_proceeds + total_cash_flow_hold - initial_investment
    roi_hold = (total_profit / initial_investment) if initial_investment > 0 else None

    hold_irr = None
    try:
        hold_irr = npf.irr(cash_flows_hold)
        if hold_irr is None or np.isnan(hold_irr):
            hold_irr = _irr_stable(np.array(cash_flows_hold))
    except Exception:
        hold_irr = _irr_stable(np.array(cash_flows_hold))
    hold_irr_annual = _annualize_from_monthly(hold_irr)

    metrics = {
        "Loan Amount": loan_amount,
        "Monthly Payment (Note Rate)": monthly_note_payment,
        "Borrower Payment (Yr1)": borrower_pmt_y1 if use_buydown_2_1 else monthly_note_payment,
        "Borrower Payment (Yr2)": borrower_pmt_y2 if use_buydown_2_1 else monthly_note_payment,
        "Effective Rate (Yr1)": eff_rate_y1 if use_buydown_2_1 else annual_rate,
        "Effective Rate (Yr2)": eff_rate_y2 if use_buydown_2_1 else annual_rate,
        "Buydown Subsidy (Yr1)": subsidy_y1 if use_buydown_2_1 else 0.0,
        "Buydown Subsidy (Yr2)": subsidy_y2 if use_buydown_2_1 else 0.0,
        "Buydown Subsidy (Total 2 yrs)": subsidy_total if use_buydown_2_1 else 0.0,
        "Buydown Paid By": buydown_paid_by if use_buydown_2_1 else "N/A",
        "NOI (annual, Yr1 levels)": noi,
        "Annual Debt Service (Yr1)": annual_debt_y1,
        "Annual Cash Flow (Yr1)": annual_cash_flow_y1,
        "Annual Debt Service (Yr2)": annual_debt_y2 if nper > 12 else None,
        "Annual Cash Flow (Yr2)": annual_cash_flow_y2 if nper > 12 else None,
        "Cap Rate": (noi / purchase_price) if purchase_price else None,
        "Cash on Cash Return (Yr1)": (annual_cash_flow_y1 / initial_investment) if initial_investment > 0 else None,
        "DCR (Yr1)": dcr_y1,
        "Break-Even Occupancy % (Yr1)": break_even_occ,    # fraction (0-1+)
        "Annual IRR (term sale)": annual_irr_term,
        "Hold Period (years)": hold_years,
        "Sale Price (hold)": sale_price,
        "Loan Balance at Hold": loan_balance_at_hold,
        "Net Sale Proceeds": net_sale_proceeds,
        "Total Cash Flow (hold)": total_cash_flow_hold,
        "Equity at Hold": equity_at_hold,
        "Total Profit (hold)": total_profit,
        "ROI (hold period)": roi_hold,
        "IRR (hold period)": hold_irr_annual,
        "Initial Investment (cash)": initial_investment,
        "Buy Closing Costs Paid By": buy_cc_paid_by,
    }
    return metrics, amort, yearly_cf_df

# -------------------------
# Excel export
# -------------------------
def dict_to_df_rowwise(d: dict, title: str):
    return pd.DataFrame(list(d.items()), columns=[title, "Value"])

def dataframes_to_xlsx_bytes(inputs: dict, metrics: dict, amort: pd.DataFrame, yearly_cf: pd.DataFrame, dscr_df: pd.DataFrame):
    tmp = BytesIO()
    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        dict_to_df_rowwise(inputs, "Input").to_excel(writer, sheet_name="Inputs", index=False)
        dict_to_df_rowwise(metrics, "Metric").to_excel(writer, sheet_name="Metrics", index=False)
        amort.to_excel(writer, sheet_name="Amortization", index=False)
        yearly_cf.to_excel(writer, sheet_name="Yearly Cashflows", index=False)
        dscr_df.to_excel(writer, sheet_name="DSCR Trend", index=False)
    tmp.seek(0)

    wb = load_workbook(tmp)

    currency_fmt = '"$"#,##0.00'
    percent_fmt  = '0.00%'
    number2_fmt  = '#,##0.00'

    # Red negatives across all sheets
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = Font(color="FF0000")

    # Inputs sheet number formats
    try:
        ws_in = wb["Inputs"]
        for r in range(2, ws_in.max_row + 1):
            label = str(ws_in.cell(row=r, column=1).value or "")
            vcell = ws_in.cell(row=r, column=2)
            val = vcell.value
            if not isinstance(val, (int, float)):
                continue
            if "%" in label:
                vcell.number_format = percent_fmt
            elif any(k in label for k in ["Price", "Payment", "Closing Costs", "Insurance", "Taxes",
                                          "HOA", "Rent", "Commission", "Monthly", "Sale Price", "Down Payment $"]):
                vcell.number_format = currency_fmt
            else:
                vcell.number_format = number2_fmt
    except Exception:
        pass

    # Metrics sheet: number formats + DCR/Break-even styling
    try:
        ws = wb["Metrics"]
        percent_keys = ["Cap Rate", "Cash on Cash", "IRR", "ROI", "Effective Rate", "Occupancy"]
        currency_keys = ["Price", "Payment", "Subsidy", "NOI", "Debt Service", "Cash Flow",
                         "Loan", "Proceeds", "Equity", "Profit", "Investment", "Closing Costs", "Commission", "Balance"]

        for r in range(2, ws.max_row + 1):
            mcell = ws.cell(row=r, column=1)
            vcell = ws.cell(row=r, column=2)
            metric_name = str(mcell.value) if mcell.value is not None else ""
            val = vcell.value

            if isinstance(val, (int, float)):
                if any(k in metric_name for k in percent_keys):
                    vcell.number_format = percent_fmt
                elif metric_name.startswith("DCR"):
                    vcell.number_format = number2_fmt
                elif any(k in metric_name for k in currency_keys):
                    vcell.number_format = currency_fmt
                else:
                    vcell.number_format = number2_fmt

            # Bold & conditional styling
            if metric_name.startswith("DCR"):
                mcell.font = Font(bold=True)
                vcell.font = Font(bold=True)
                if isinstance(val, (int, float)):
                    if val < 1.0:
                        vcell.font = Font(bold=True, color="FF0000")
                    elif val >= 1.25:
                        vcell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            if metric_name.startswith("Break-Even Occupancy"):
                mcell.font = Font(bold=True)
                vcell.font = Font(bold=True)
                vcell.number_format = percent_fmt
                if isinstance(val, (int, float)) and val > 1.0:
                    vcell.font = Font(bold=True, color="FF0000")
    except Exception:
        pass

    # Yearly Cashflows sheet formats
    try:
        ws_cf = wb["Yearly Cashflows"]
        money_cols = [
            "Gross Rent (after vacancy) /mo", "Mgmt Fee /mo",
            "Other OpEx /mo (excl taxes/ins/HOA)", "Taxes /mo", "Insurance /mo", "HOA /mo",
            "Total OpEx /mo", "NOI (annual, prorated)",
            "Debt Service (annual, prorated)", "Cash Flow (annual)", "Cumulative Cash Flow"
        ]
        headers = {ws_cf.cell(row=1, column=c).value: c for c in range(1, ws_cf.max_column + 1)}
        for name in money_cols:
            if name in headers:
                col = headers[name]
                for r in range(2, ws_cf.max_row + 1):
                    cell = ws_cf.cell(row=r, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = currency_fmt
        for name in ["Year", "Months Counted"]:
            if name in headers:
                col = headers[name]
                for r in range(2, ws_cf.max_row + 1):
                    cell = ws_cf.cell(row=r, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
    except Exception:
        pass

    # Amortization sheet formats
    try:
        ws_am = wb["Amortization"]
        headers = {ws_am.cell(row=1, column=c).value: c for c in range(1, ws_am.max_column + 1)}
        money_cols = ["Note Payment", "Borrower Payment", "Subsidy", "Principal", "Interest", "Balance", "Equity"]
        for name in money_cols:
            if name in headers:
                col = headers[name]
                for r in range(2, ws_am.max_row + 1):
                    cell = ws_am.cell(row=r, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = currency_fmt
        if "Month" in headers:
            col = headers["Month"]
            for r in range(2, ws_am.max_row + 1):
                cell = ws_am.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0'
    except Exception:
        pass

    # DSCR Trend sheet formats + conditional styling
    try:
        ws_d = wb["DSCR Trend"]
        headers = {ws_d.cell(row=1, column=c).value: c for c in range(1, ws_d.max_column + 1)}
        if "NOI (annual, prorated)" in headers:
            for r in range(2, ws_d.max_row + 1):
                cell = ws_d.cell(row=r, column=headers["NOI (annual, prorated)"])
                if isinstance(cell.value, (int, float)): cell.number_format = currency_fmt
        if "Debt Service (annual, prorated)" in headers:
            for r in range(2, ws_d.max_row + 1):
                cell = ws_d.cell(row=r, column=headers["Debt Service (annual, prorated)"])
                if isinstance(cell.value, (int, float)): cell.number_format = currency_fmt
        if "DCR" in headers:
            for r in range(2, ws_d.max_row + 1):
                cell = ws_d.cell(row=r, column=headers["DCR"])
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number2_fmt
                    if cell.value < 1.0:
                        cell.font = Font(color="FF0000", bold=True)
                    elif cell.value >= 1.25:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True)
        if "Year" in headers:
            for r in range(2, ws_d.max_row + 1):
                cell = ws_d.cell(row=r, column=headers["Year"])
                if isinstance(cell.value, (int, float)): cell.number_format = '0'
        for c in range(1, ws_d.max_column + 1):
            ws_d.cell(row=1, column=c).font = Font(bold=True)
    except Exception:
        pass

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# -------------------------
# Lead capture helpers
# -------------------------
def _is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", (email or "").strip()))


def _append_local_csv(row: dict, file_name: str = "leads.csv"):
    """Fallback storage for local testing. On Streamlit Cloud this may not be permanent."""
    path = Path(file_name)
    df = pd.DataFrame([row])
    if path.exists():
        df.to_csv(path, mode="a", header=False, index=False)
    else:
        df.to_csv(path, index=False)


def _append_google_sheet(row: dict, worksheet_name: str = "Leads"):
    """
    Optional durable lead storage using Streamlit secrets.

    Add these secrets in Streamlit Cloud:
    [google_service_account]
    type = "service_account"
    project_id = "..."
    private_key_id = "..."
    private_key = "-----BEGIN PRIVATE KEY-----\\n...\\n-----END PRIVATE KEY-----\\n"
    client_email = "..."
    client_id = "..."
    auth_uri = "https://accounts.google.com/o/oauth2/auth"
    token_uri = "https://oauth2.googleapis.com/token"
    auth_provider_x509_cert_url = "https://www.googleapis.com/oauth2/v1/certs"
    client_x509_cert_url = "..."

    [google_sheet]
    sheet_id = "YOUR_GOOGLE_SHEET_ID"
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        sheet_id = st.secrets.get("google_sheet", {}).get("sheet_id", "")
        service_account_info = st.secrets.get("google_service_account", None)
        if not sheet_id or not service_account_info:
            return False, "Google Sheets secrets not configured. Saved locally instead."

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scopes)
        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_key(sheet_id)
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except Exception:
            worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=30)

        existing = worksheet.get_all_values()
        headers = list(row.keys())
        if not existing:
            worksheet.append_row(headers)
        worksheet.append_row([row.get(h, "") for h in headers])
        return True, "Saved to Google Sheets."
    except Exception as e:
        return False, f"Google Sheets save failed: {e}. Saved locally instead."


def _save_lead(row: dict):
    ok, msg = _append_google_sheet(row, worksheet_name="Leads")
    if not ok:
        _append_local_csv(row, "leads.csv")
    return ok, msg


def _save_feedback(row: dict):
    ok, msg = _append_google_sheet(row, worksheet_name="Feedback")
    if not ok:
        _append_local_csv(row, "feedback.csv")
    return ok, msg


def _get_session_id():
    """Create one anonymous session id per browser session."""
    if "session_id" not in st.session_state:
        st.session_state.session_id = str(uuid.uuid4())
    return st.session_state.session_id


def _track_usage(event: str, extra: dict | None = None):
    """Track app usage events to Google Sheets, with local CSV fallback.

    Events include: page_visit, lead_submitted, calculation_run, feedback, excel_download.
    Local CSV is fine for testing, but Streamlit Cloud storage is not permanent.
    Use Google Sheets secrets for durable tracking.
    """
    lead = st.session_state.get("lead_info", {})
    row = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "event": event,
        "session_id": _get_session_id(),
        "name": lead.get("name", ""),
        "email": lead.get("email", ""),
        "phone": lead.get("phone", ""),
        "investor_type": lead.get("investor_type", ""),
        "strategy": lead.get("strategy", ""),
        "market": lead.get("market", ""),
        "budget": lead.get("budget", ""),
        "source": "Smart Rental ROI App",
    }
    if extra:
        row.update(extra)

    ok, msg = _append_google_sheet(row, worksheet_name="Usage")
    if not ok:
        _append_local_csv(row, "usage_tracking.csv")
    return ok, msg


def _track_page_visit_once():
    if "page_visit_tracked" not in st.session_state:
        st.session_state.page_visit_tracked = True
        _track_usage("page_visit")


def _admin_usage_dashboard():
    """Optional mini dashboard for you. Set secrets admin.password to enable."""
    with st.sidebar.expander("Admin Usage", expanded=False):
        configured_password = st.secrets.get("admin", {}).get("password", "") if hasattr(st, "secrets") else ""
        password = st.text_input("Admin password", type="password")
        if not configured_password:
            st.caption("Set [admin] password in Streamlit secrets to enable this dashboard.")
            return
        if password != configured_password:
            return

        usage_path = Path("usage_tracking.csv")
        leads_path = Path("leads.csv")
        if usage_path.exists():
            usage_df = pd.read_csv(usage_path)
            st.metric("Tracked Events", len(usage_df))
            if "event" in usage_df.columns:
                st.write(usage_df["event"].value_counts())
            st.dataframe(usage_df.tail(25), use_container_width=True)
        else:
            st.info("No local usage CSV found. If Google Sheets is configured, check the Usage worksheet.")

        if leads_path.exists():
            leads_df = pd.read_csv(leads_path)
            st.metric("Local Leads", len(leads_df))


def _lead_capture_gate():
    st.markdown("### Get Your Free Investment Analysis")
    st.caption("Enter your contact details to use the ROI calculator and download the Excel report.")

    if "lead_captured" not in st.session_state:
        st.session_state.lead_captured = False
    if "lead_info" not in st.session_state:
        st.session_state.lead_info = {}

    if st.session_state.lead_captured:
        lead = st.session_state.lead_info
        st.success(f"Welcome {lead.get('name', '')}! You can run the calculator below.")
        return

    with st.form("lead_capture_form", clear_on_submit=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            name = st.text_input("Name *")
            email = st.text_input("Email *")
        with c2:
            phone = st.text_input("Phone")
            investor_type = st.selectbox("I am a", ["Investor", "Agent", "Buyer", "Seller", "Wholesaler", "Lender", "Other"])
        with c3:
            strategy = st.selectbox("Investment Strategy", ["Buy & Hold", "BRRRR", "Flip", "Multifamily", "Short-Term Rental", "Commercial", "Just researching"])
            market = st.text_input("Preferred Market", value="Charlotte / NC / SC")

        budget = st.selectbox("Purchase Budget", ["Under $250K", "$250K-$500K", "$500K-$1M", "$1M+", "Not sure"])
        property_address = st.text_input("Property you want to analyze (optional)")
        consent = st.checkbox("I agree to be contacted about this calculator, deal feedback, and investment opportunities.")
        submitted_lead = st.form_submit_button("Start ROI Analysis", use_container_width=True)

    if submitted_lead:
        if not name.strip():
            st.error("Please enter your name.")
            st.stop()
        if not _is_valid_email(email):
            st.error("Please enter a valid email.")
            st.stop()
        if not consent:
            st.error("Please check the contact permission box to continue.")
            st.stop()

        row = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "name": name.strip(),
            "email": email.strip(),
            "phone": phone.strip(),
            "investor_type": investor_type,
            "strategy": strategy,
            "market": market.strip(),
            "budget": budget,
            "property_address": property_address.strip(),
            "source": "Smart Rental ROI App",
        }
        _save_lead(row)
        st.session_state.lead_info = row
        st.session_state.lead_captured = True
        _track_usage("lead_submitted", {
            "property_address": property_address.strip(),
            "consent": consent,
        })
        st.rerun()

    st.stop()



def _investor_deals_list_section(default_purchase_price=0.0, default_market="Charlotte / NC / SC"):
    """Soft lead capture: no signup required to use calculator."""
    st.divider()
    st.subheader("🔥 Get Weekly Charlotte Investor Deals")
    st.markdown(
        "Join the investor list to receive off-market leads, flips, BRRRR opportunities, "
        "multifamily alerts, price reductions, and deal-analysis tips."
    )

    with st.expander("Join Investor Deals List", expanded=False):
        with st.form("investor_deals_list_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                investor_name = st.text_input("Name *", key="deals_name")
                investor_email = st.text_input("Email *", key="deals_email")
                investor_phone = st.text_input("Phone", key="deals_phone")
            with c2:
                investor_type = st.selectbox(
                    "I am a",
                    ["Investor", "Agent", "Buyer", "Seller", "Wholesaler", "Lender", "Other"],
                    key="deals_investor_type"
                )
                budget = st.selectbox(
                    "Purchase Budget",
                    ["Under $250K", "$250K-$500K", "$500K-$1M", "$1M+", "Not sure"],
                    key="deals_budget"
                )
                timeline = st.selectbox(
                    "Buying Timeline",
                    ["Now", "0-3 months", "3-6 months", "6+ months", "Just researching"],
                    key="deals_timeline"
                )

            strategy = st.multiselect(
                "Investment Strategy",
                ["Buy & Hold", "BRRRR", "Flip", "Multifamily", "Short-Term Rental", "Commercial", "Land"],
                default=["Buy & Hold"],
                key="deals_strategy"
            )
            market = st.multiselect(
                "Preferred Markets",
                ["Charlotte", "Waxhaw", "Fort Mill", "Indian Trail", "Concord", "Denver", "Hickory", "Rock Hill", "Gastonia", "Kannapolis", "Other NC/SC"],
                default=["Charlotte"],
                key="deals_market"
            )
            notes = st.text_area("What type of deals are you looking for?", key="deals_notes")
            consent = st.checkbox(
                "I agree to be contacted about investor deals, property analysis, and market updates.",
                key="deals_consent"
            )
            submitted = st.form_submit_button("Join Investor Deals List", use_container_width=True)

        if submitted:
            if not investor_name.strip():
                st.error("Please enter your name.")
                return
            if not _is_valid_email(investor_email):
                st.error("Please enter a valid email.")
                return
            if not consent:
                st.error("Please check the contact permission box.")
                return

            row = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "name": investor_name.strip(),
                "email": investor_email.strip(),
                "phone": investor_phone.strip(),
                "investor_type": investor_type,
                "strategy": ", ".join(strategy),
                "market": ", ".join(market) if market else default_market,
                "budget": budget,
                "timeline": timeline,
                "notes": notes.strip(),
                "current_purchase_price": default_purchase_price,
                "source": "Investor Deals List - Smart Rental ROI App",
                "consent": consent,
            }
            _save_lead(row)
            st.session_state.lead_info = row
            st.session_state.lead_captured = True
            _track_usage("investor_deals_signup", {
                "budget": budget,
                "strategy": row["strategy"],
                "market": row["market"],
                "timeline": timeline,
                "current_purchase_price": default_purchase_price,
            })
            st.success("Thank you! You are on the investor deals list.")




def _lead_score(actions: dict) -> tuple[int, str]:
    """Simple lead scoring for follow-up priority."""
    score = 0
    score += 5 if actions.get("visit") else 0
    score += 10 if actions.get("calculation") else 0
    score += 20 if actions.get("excel") else 0
    score += 25 if actions.get("investor_list") else 0
    score += 30 if actions.get("email_report") else 0
    score += 35 if actions.get("loan_quote") else 0
    score += 40 if actions.get("seller_lead") else 0
    score += 50 if actions.get("deal_review") else 0
    score += 30 if actions.get("buy_box") else 0
    if score >= 75:
        return score, "Hot Lead"
    if score >= 35:
        return score, "Warm Lead"
    return score, "Cold Lead"


def _current_metrics_summary(metrics: dict, purchase_price: float, monthly_rent: float, num_units: int) -> dict:
    return {
        "purchase_price": purchase_price,
        "monthly_rent_total": monthly_rent,
        "number_of_units": num_units,
        "cap_rate": metrics.get("Cap Rate"),
        "cash_on_cash": metrics.get("Cash on Cash Return (Yr1)"),
        "dcr": metrics.get("DCR (Yr1)"),
        "irr_hold": metrics.get("IRR (hold period)"),
        "roi_hold": metrics.get("ROI (hold period)"),
        "total_profit_hold": metrics.get("Total Profit (hold)"),
        "annual_cash_flow_y1": metrics.get("Annual Cash Flow (Yr1)"),
        "net_sale_proceeds": metrics.get("Net Sale Proceeds"),
    }


def _fmt_pct_for_text(x):
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return "N/A"
        return f"{x*100:.2f}%"
    except Exception:
        return "N/A"


def _fmt_money_for_text(x):
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return "N/A"
        return f"${x:,.2f}"
    except Exception:
        return "N/A"


def _lead_common_fields(name, email, phone, lead_type, source, extra=None):
    extra = extra or {}
    actions = {
        "visit": True,
        "calculation": True,
        "excel": lead_type == "Excel Download",
        "investor_list": lead_type == "Investor Deals List",
        "email_report": lead_type == "Email ROI Report",
        "deal_review": lead_type == "Deal Review Request",
        "buy_box": lead_type == "Investor Buy Box",
        "loan_quote": lead_type == "DSCR Loan Quote",
        "seller_lead": lead_type == "Seller Lead",
    }
    score, rating = _lead_score(actions)
    row = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "lead_type": lead_type,
        "name": str(name or "").strip(),
        "email": str(email or "").strip(),
        "phone": str(phone or "").strip(),
        "lead_score": score,
        "lead_rating": rating,
        "source": source,
        "session_id": _get_session_id(),
    }
    row.update(extra)
    return row


def _email_roi_report_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("📧 Email Me This ROI Report")
    st.caption("Let users run the analysis first, then capture contact info from serious users who want the report sent to them.")
    with st.expander("Email this analysis", expanded=False):
        with st.form("email_roi_report_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                name = st.text_input("Name *", key="email_report_name")
            with c2:
                email = st.text_input("Email *", key="email_report_email")
            with c3:
                phone = st.text_input("Phone", key="email_report_phone")
            property_address = st.text_input("Property Address / MLS #", key="email_report_property")
            consent = st.checkbox("I agree to be contacted about this ROI report and investment opportunities.", key="email_report_consent")
            submitted = st.form_submit_button("Request ROI Report", use_container_width=True)
        if submitted:
            if not name.strip():
                st.error("Please enter your name.")
                return
            if not _is_valid_email(email):
                st.error("Please enter a valid email.")
                return
            if not consent:
                st.error("Please check the contact permission box.")
                return
            extra = _current_metrics_summary(metrics, purchase_price, monthly_rent, num_units)
            extra.update({"property_address": property_address.strip(), "consent": consent})
            row = _lead_common_fields(name, email, phone, "Email ROI Report", "Smart Rental ROI App", extra)
            _save_lead(row)
            st.session_state.lead_info = row
            _track_usage("email_roi_report_request", extra)
            st.success("Request saved. You can follow up with this user and send the report.")


def _deal_review_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("📋 Request Free Deal Review")
    st.caption("Highest-quality lead: this user has a specific property and wants your second opinion.")
    with st.expander("Request free deal review", expanded=False):
        with st.form("deal_review_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("Name *", key="review_name")
                email = st.text_input("Email *", key="review_email")
                phone = st.text_input("Phone", key="review_phone")
            with c2:
                property_address = st.text_input("Property Address / MLS # *", key="review_property")
                timeline = st.selectbox("Decision Timeline", ["Now", "This week", "This month", "Just researching"], key="review_timeline")
                financing = st.selectbox("Financing", ["Cash", "Conventional", "DSCR Loan", "Hard Money", "Not sure"], key="review_financing")
            notes = st.text_area("Questions / concerns about the deal", key="review_notes")
            consent = st.checkbox("I agree to be contacted about this deal review.", key="review_consent")
            submitted = st.form_submit_button("Request Free Deal Review", use_container_width=True)
        if submitted:
            if not name.strip() or not property_address.strip():
                st.error("Please enter your name and property address.")
                return
            if not _is_valid_email(email):
                st.error("Please enter a valid email.")
                return
            if not consent:
                st.error("Please check the contact permission box.")
                return
            extra = _current_metrics_summary(metrics, purchase_price, monthly_rent, num_units)
            extra.update({"property_address": property_address.strip(), "timeline": timeline, "financing": financing, "notes": notes.strip(), "consent": consent})
            row = _lead_common_fields(name, email, phone, "Deal Review Request", "Smart Rental ROI App", extra)
            _save_lead(row)
            st.session_state.lead_info = row
            _track_usage("deal_review_request", extra)
            st.success("Deal review request saved.")


def _investor_buy_box_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("🎯 Investor Buy Box")
    st.caption("Use this to learn exactly what each investor wants to buy.")
    with st.expander("Build my investor buy box", expanded=False):
        with st.form("investor_buy_box_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("Name *", key="buybox_name")
                email = st.text_input("Email *", key="buybox_email")
                phone = st.text_input("Phone", key="buybox_phone")
                budget = st.selectbox("Budget", ["Under $250K", "$250K-$500K", "$500K-$1M", "$1M+"], key="buybox_budget")
            with c2:
                strategy = st.multiselect("Strategy", ["Buy & Hold", "BRRRR", "Flip", "Multifamily", "Short-Term Rental", "Commercial", "Land"], key="buybox_strategy")
                property_type = st.multiselect("Property Type", ["Single Family", "Townhome", "Condo", "Duplex", "Triplex/Quad", "5+ Unit Multifamily", "Commercial"], key="buybox_property_type")
                timeline = st.selectbox("Timeline", ["Now", "0-3 months", "3-6 months", "6+ months"], key="buybox_timeline")
                financing = st.selectbox("Financing", ["Cash", "Conventional", "DSCR Loan", "Hard Money", "Private Money", "Not sure"], key="buybox_financing")
            markets = st.text_input("Preferred markets", value="Charlotte, Fort Mill, Waxhaw, Hickory, Denver", key="buybox_markets")
            criteria = st.text_area("Must-have criteria", key="buybox_criteria")
            consent = st.checkbox("I agree to be contacted with matching investor opportunities.", key="buybox_consent")
            submitted = st.form_submit_button("Save My Buy Box", use_container_width=True)
        if submitted:
            if not name.strip():
                st.error("Please enter your name.")
                return
            if not _is_valid_email(email):
                st.error("Please enter a valid email.")
                return
            if not consent:
                st.error("Please check the contact permission box.")
                return
            extra = _current_metrics_summary(metrics, purchase_price, monthly_rent, num_units)
            extra.update({
                "budget": budget,
                "strategy": ", ".join(strategy),
                "property_type": ", ".join(property_type),
                "timeline": timeline,
                "financing": financing,
                "market": markets.strip(),
                "criteria": criteria.strip(),
                "consent": consent,
            })
            row = _lead_common_fields(name, email, phone, "Investor Buy Box", "Smart Rental ROI App", extra)
            _save_lead(row)
            st.session_state.lead_info = row
            _track_usage("investor_buy_box", extra)
            st.success("Your buy box was saved.")


def _save_compare_scenario_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("💾 Save & Compare Scenarios")
    if "saved_scenarios" not in st.session_state:
        st.session_state.saved_scenarios = []
    with st.expander("Save current scenario", expanded=False):
        scenario_name = st.text_input("Scenario Name", value=f"Property {len(st.session_state.saved_scenarios)+1}", key="scenario_name")
        if st.button("Save Scenario", use_container_width=True):
            row = {
                "Scenario": scenario_name,
                "Purchase Price": purchase_price,
                "Monthly Rent": monthly_rent,
                "Units": num_units,
                "Annual Cash Flow Yr1": metrics.get("Annual Cash Flow (Yr1)"),
                "Cap Rate": metrics.get("Cap Rate"),
                "Cash on Cash": metrics.get("Cash on Cash Return (Yr1)"),
                "DCR": metrics.get("DCR (Yr1)"),
                "IRR Hold": metrics.get("IRR (hold period)"),
                "Total Profit Hold": metrics.get("Total Profit (hold)"),
            }
            st.session_state.saved_scenarios.append(row)
            _track_usage("scenario_saved", {"scenario_name": scenario_name, "purchase_price": purchase_price})
            st.success("Scenario saved for this session.")
    if st.session_state.saved_scenarios:
        compare_df = pd.DataFrame(st.session_state.saved_scenarios)
        fmt = {
            "Purchase Price": "${:,.0f}",
            "Monthly Rent": "${:,.0f}",
            "Annual Cash Flow Yr1": "${:,.0f}",
            "Cap Rate": "{:.2%}",
            "Cash on Cash": "{:.2%}",
            "DCR": "{:.2f}",
            "IRR Hold": "{:.2%}",
            "Total Profit Hold": "${:,.0f}",
        }
        st.dataframe(compare_df.style.format({k:v for k,v in fmt.items() if k in compare_df.columns}), use_container_width=True)
        st.download_button("Download Scenario Comparison CSV", compare_df.to_csv(index=False), "scenario_comparison.csv", "text/csv", use_container_width=True)


def _share_analysis_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("🔗 Share Analysis")
    summary_text = f"""Smart Rental ROI Analysis
Purchase Price: {_fmt_money_for_text(purchase_price)}
Monthly Rent: {_fmt_money_for_text(monthly_rent)}
Units: {num_units}
Annual Cash Flow Yr1: {_fmt_money_for_text(metrics.get('Annual Cash Flow (Yr1)'))}
Cap Rate: {_fmt_pct_for_text(metrics.get('Cap Rate'))}
Cash-on-Cash Return Yr1: {_fmt_pct_for_text(metrics.get('Cash on Cash Return (Yr1)'))}
DCR Yr1: {metrics.get('DCR (Yr1)') if metrics.get('DCR (Yr1)') is not None else 'N/A'}
IRR Hold Period: {_fmt_pct_for_text(metrics.get('IRR (hold period)'))}
Total Profit Hold: {_fmt_money_for_text(metrics.get('Total Profit (hold)'))}
"""
    with st.expander("Copy/share result summary", expanded=False):
        st.text_area("Copy this summary to send to a partner, lender, or client", summary_text, height=220)
        _track_usage("share_summary_viewed", {"purchase_price": purchase_price})


def _lead_score_section():
    lead = st.session_state.get("lead_info", {})
    if lead:
        score = lead.get("lead_score", "")
        rating = lead.get("lead_rating", "")
        if score != "":
            st.info(f"Lead status for latest opt-in: {rating} — Score {score}")


def _loan_quote_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("🏦 Need DSCR / Investor Loan Quote?")
    with st.expander("Request loan quote", expanded=False):
        with st.form("loan_quote_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("Name *", key="loan_name")
                email = st.text_input("Email *", key="loan_email")
                phone = st.text_input("Phone", key="loan_phone")
                loan_amount = st.number_input("Estimated Loan Amount ($)", min_value=0.0, value=float(max(0, purchase_price * 0.8)), step=5000.0, key="loan_amount_quote")
            with c2:
                credit_range = st.selectbox("Credit Score Range", ["760+", "720-759", "680-719", "640-679", "Below 640", "Not sure"], key="loan_credit")
                down_payment_quote = st.number_input("Down Payment Available ($)", min_value=0.0, value=float(max(0, purchase_price * 0.2)), step=5000.0, key="loan_down_payment")
                property_address = st.text_input("Property Address / MLS #", key="loan_property")
                loan_type = st.selectbox("Loan Type", ["DSCR", "Conventional Investor", "Hard Money", "Bridge", "Not sure"], key="loan_type")
            consent = st.checkbox("I agree to be contacted about financing options.", key="loan_consent")
            submitted = st.form_submit_button("Request Loan Quote", use_container_width=True)
        if submitted:
            if not name.strip():
                st.error("Please enter your name.")
                return
            if not _is_valid_email(email):
                st.error("Please enter a valid email.")
                return
            if not consent:
                st.error("Please check the contact permission box.")
                return
            extra = _current_metrics_summary(metrics, purchase_price, monthly_rent, num_units)
            extra.update({"loan_amount_requested": loan_amount, "credit_range": credit_range, "down_payment_available": down_payment_quote, "property_address": property_address.strip(), "loan_type": loan_type, "consent": consent})
            row = _lead_common_fields(name, email, phone, "DSCR Loan Quote", "Smart Rental ROI App", extra)
            _save_lead(row)
            st.session_state.lead_info = row
            _track_usage("loan_quote_request", extra)
            st.success("Loan quote request saved.")


def _seller_lead_section(metrics, purchase_price, monthly_rent, num_units):
    st.subheader("🏠 Have a Property to Sell?")
    with st.expander("Get investor offer estimate", expanded=False):
        with st.form("seller_lead_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("Name *", key="seller_name")
                email = st.text_input("Email *", key="seller_email")
                phone = st.text_input("Phone", key="seller_phone")
                property_address = st.text_input("Property Address *", key="seller_property")
            with c2:
                asking_price = st.number_input("Target / Asking Price ($)", min_value=0.0, value=float(purchase_price), step=5000.0, key="seller_asking")
                condition = st.selectbox("Condition", ["Move-in ready", "Light updates", "Needs rehab", "Heavy rehab", "Not sure"], key="seller_condition")
                timeline = st.selectbox("Selling Timeline", ["ASAP", "0-30 days", "30-90 days", "3+ months", "Just curious"], key="seller_timeline")
                occupancy = st.selectbox("Occupancy", ["Vacant", "Owner occupied", "Tenant occupied", "Partially occupied", "Not sure"], key="seller_occupancy")
            notes = st.text_area("Notes", key="seller_notes")
            consent = st.checkbox("I agree to be contacted about selling this property.", key="seller_consent")
            submitted = st.form_submit_button("Request Investor Offer Estimate", use_container_width=True)
        if submitted:
            if not name.strip() or not property_address.strip():
                st.error("Please enter your name and property address.")
                return
            if not _is_valid_email(email):
                st.error("Please enter a valid email.")
                return
            if not consent:
                st.error("Please check the contact permission box.")
                return
            extra = _current_metrics_summary(metrics, purchase_price, monthly_rent, num_units)
            extra.update({"property_address": property_address.strip(), "asking_price": asking_price, "condition": condition, "timeline": timeline, "occupancy": occupancy, "notes": notes.strip(), "consent": consent})
            row = _lead_common_fields(name, email, phone, "Seller Lead", "Smart Rental ROI App", extra)
            _save_lead(row)
            st.session_state.lead_info = row
            _track_usage("seller_lead_request", extra)
            st.success("Seller lead request saved.")

# -------------------------
# Streamlit UI
# -------------------------
_track_page_visit_once()
_admin_usage_dashboard()

# Do NOT require lead info upfront. Let users run the calculator anonymously.
# Lead capture happens later via the Investor Deals List, Feedback, Deal Review,
# and Excel download events.
if "lead_info" not in st.session_state:
    st.session_state.lead_info = {}
if "lead_captured" not in st.session_state:
    st.session_state.lead_captured = False

st.divider()

col1, col2, col3 = st.columns([1, 1, 1])

with col1:
    dp_mode = st.radio("Down Payment Mode", ["Percent of Price", "Dollar Amount"], horizontal=True)
    purchase_price = st.number_input("Purchase Price ($)", value=300000.0, min_value=0.0, step=1000.0)
    if dp_mode == "Percent of Price":
        down_payment_pct = st.number_input("Down Payment (% of Price)", value=20.0, min_value=0.0, max_value=100.0, step=1.0)
        down_payment = purchase_price * (down_payment_pct / 100.0)
    else:
        down_payment = st.number_input("Down Payment ($)", value=60000.0, min_value=0.0, step=1000.0)
        down_payment_pct = (down_payment / purchase_price * 100.0) if purchase_price > 0 else 0.0
    st.markdown(f"**Down Payment:** ${down_payment:,.2f}  •  **{down_payment_pct:.2f}%**")
    st.markdown(f"**Loan Amount:** ${purchase_price - down_payment:,.2f}")

with col2:
    interest_rate_pct = st.number_input("Interest Rate (annual %)", value=6.5, min_value=0.0, step=0.01)
    loan_term_years = st.number_input("Loan Term (years)", value=30, min_value=0, step=1)
    override_payment = st.number_input("Monthly Loan Payment (optional)", value=0.0, min_value=0.0, step=50.0)
    use_buydown_2_1 = st.checkbox("Use 2/1 Buydown (Yr1 = note%-2, Yr2 = note%-1)")
    if use_buydown_2_1:
        buydown_paid_by = st.radio("Buydown escrow paid by", ["Buyer", "Seller", "Lender Credit"], horizontal=True)
    else:
        buydown_paid_by = "Buyer"

with col3:
    freq = st.radio("Rent & Expenses Frequency (excluding taxes/insurance)", ["Monthly", "Yearly"], horizontal=True)

# --- Buy-side closing costs payer ---
buy_cc_paid_by = st.radio("Buy-side Closing Costs Paid By", ["Buyer", "Seller", "Lender Credit"], horizontal=True)

# --- Units & rents ---
st.subheader("🏘️ Units")
num_units = st.number_input("Number of Units", min_value=1, value=1, step=1)
unit_rents = []
for i in range(int(num_units)):
    if freq == "Monthly":
        rent = st.number_input(f"Unit {i+1} Rent ($/month)", value=1000.0, min_value=0.0, step=50.0, key=f"unit_{i}_rent_monthly")
    else:
        rent = st.number_input(f"Unit {i+1} Rent ($/year)", value=12000.0, min_value=0.0, step=500.0, key=f"unit_{i}_rent_yearly") / 12.0
    unit_rents.append(rent)
monthly_rent = float(sum(unit_rents))

if freq == "Monthly":
    monthly_expenses = st.number_input("Other Monthly Expenses ($)", value=500.0, min_value=0.0, step=50.0)
    monthly_hoa = st.number_input("Monthly HOA ($)", value=0.0, min_value=0.0, step=25.0)
else:
    yearly_expenses = st.number_input("Other Yearly Expenses ($)", value=6000.0, min_value=0.0, step=500.0)
    yearly_hoa = st.number_input("Yearly HOA ($)", value=0.0, min_value=0.0, step=100.0)
    monthly_expenses = yearly_expenses / 12.0
    monthly_hoa = yearly_hoa / 12.0

st.subheader("🏦 Annual Fixed Costs")
annual_taxes = st.number_input("Annual Property Taxes ($)", value=3600.0, min_value=0.0, step=500.0)
annual_insurance = st.number_input("Annual Insurance ($)", value=1200.0, min_value=0.0, step=100.0)
monthly_taxes = annual_taxes / 12.0
monthly_insurance = annual_insurance / 12.0

st.subheader("⚙️ Vacancy & Management")
vacancy_pct = st.number_input("Vacancy Rate (% of Gross Rent)", value=5.0, min_value=0.0, max_value=100.0, step=0.5)
mgmt_pct = st.number_input("Management Fee (% of Collected Rent)", value=8.0, min_value=0.0, max_value=100.0, step=0.5)

st.subheader("📈 Rent Growth")
rent_appreciation_pct = st.number_input("Annual Rental Appreciation Rate (%)", value=2.0, min_value=-100.0, max_value=100.0, step=0.25)

st.subheader("🏠 Appreciation & Transaction Costs")

# --- Closing Costs at Purchase ($): disabled when Seller/Lender pays ---
if "buy_cc" not in st.session_state:
    st.session_state.buy_cc = 5000.0  # editable default when Buyer pays

if buy_cc_paid_by != "Buyer":
    st.session_state.buy_cc = 0.0

buy_closing_costs = st.number_input(
    "Closing Costs at Purchase ($)",
    min_value=0.0,
    step=100.0,
    key="buy_cc",
    disabled=(buy_cc_paid_by != "Buyer")
)

sale_mode = st.radio("Sale Price Input", ["Use Annual Appreciation Rate (%)", "Enter Flat Sale Amount"])
if sale_mode == "Use Annual Appreciation Rate (%)":
    appreciation_pct = st.number_input("Annual Property Appreciation Rate (%)", value=3.0, min_value=-100.0, max_value=100.0, step=0.25)
    sale_price_override = 0.0
else:
    appreciation_pct = st.number_input("Annual Property Appreciation Rate (%)", value=0.0, min_value=-100.0, max_value=100.0, step=0.25, help="Ignored when Flat Sale Amount is provided.")
    sale_price_override = st.number_input("Flat Sale Price at Hold ($)", value=0.0, min_value=0.0, step=1000.0)

sale_closing_costs = st.number_input("Closing Costs at Sale ($)", value=5000.0, min_value=0.0, step=100.0)
sale_commission_pct = st.number_input("Sale Commission (% of Sale Price)", value=6.0, min_value=0.0, max_value=20.0, step=0.25)

st.subheader("📅 Hold Period Analysis")
hold_years = st.number_input("Hold Period (years)", value=5, min_value=1, max_value=loan_term_years if loan_term_years > 0 else 50, step=1)

# -------------------------
# Live buydown preview under the rate input
# -------------------------
loan_amount_ui = max(0.0, purchase_price - down_payment)
nper_ui = int(round(loan_term_years * 12))
note_rate_ui = interest_rate_pct / 100.0
if use_buydown_2_1 and nper_ui > 0 and loan_amount_ui > 0:
    eff1, eff2, p1, p2, note_pmt_ui, sub1, sub2, sub_total = _buydown_preview(loan_amount_ui, note_rate_ui, nper_ui)
    st.caption(f"Yr1 effective rate: **{eff1*100:.2f}%** • Borrower payment: **${p1:,.2f}**")
    st.caption(f"Yr2 effective rate: **{eff2*100:.2f}%** • Borrower payment: **${p2:,.2f}**")
    st.caption(f"Note-rate payment (amortization basis): **${note_pmt_ui:,.2f}**")
    st.info(f"Buydown escrow (lump sum at closing): **${sub_total:,.0f}**  (Yr1: ${sub1:,.0f}, Yr2: ${sub2:,.0f}).")
else:
    if nper_ui > 0 and loan_amount_ui > 0:
        note_pmt_ui = pmt(note_rate_ui/12.0, nper_ui, loan_amount_ui)
        st.caption(f"Note-rate payment: **${note_pmt_ui:,.2f}**")

# -------------------------
# Run calculation
# -------------------------
if st.button("📊 Calculate ROI", use_container_width=True):
    annual_rate = interest_rate_pct / 100.0

    # Warn on negative amortization if override_payment is too small
    if override_payment and override_payment > 0:
        first_month_interest = (annual_rate / 12.0) * max(0.0, purchase_price - down_payment)
        if override_payment + 1e-6 < first_month_interest:
            st.warning(
                f"⚠️ Your override payment (${override_payment:,.2f}) is below the first month's interest "
                f"(${first_month_interest:,.2f}). The balance will grow (negative amortization)."
            )

    metrics, amort, yearly_cf = calculate_metrics(
        purchase_price, down_payment, annual_rate, loan_term_years,
        monthly_rent, monthly_expenses, monthly_taxes, monthly_insurance, monthly_hoa,
        vacancy_pct=vacancy_pct, mgmt_pct=mgmt_pct, hold_years=hold_years,
        rent_appreciation_pct=rent_appreciation_pct,
        appreciation_pct=appreciation_pct, buy_closing_costs=buy_closing_costs,
        sale_closing_costs=sale_closing_costs, sale_commission_pct=sale_commission_pct,
        override_payment=override_payment, use_buydown_2_1=use_buydown_2_1,
        buydown_paid_by=buydown_paid_by, sale_price_override=sale_price_override,
        buy_cc_paid_by=buy_cc_paid_by
    )

    inputs = {
        "Purchase Price": purchase_price,
        "Down Payment %": down_payment_pct / 100.0,
        "Down Payment $": down_payment,
        "Loan Term (years)": loan_term_years,
        "Interest Rate % (note)": interest_rate_pct / 100.0,
        "Use 2/1 Buydown": use_buydown_2_1,
        "Buydown Paid By": buydown_paid_by if use_buydown_2_1 else "N/A",
        "Buy-side CC Paid By": buy_cc_paid_by,
        "Closing Costs (Buy)": buy_closing_costs,
        "Override Payment": override_payment,
        "Annual Taxes": annual_taxes,
        "Annual Insurance": annual_insurance,
        "Other Monthly Expenses": monthly_expenses,
        "Monthly HOA": monthly_hoa,
        "Vacancy Rate %": vacancy_pct / 100.0,
        "Management Fee %": mgmt_pct / 100.0,
        "Annual Rent Appreciation %": rent_appreciation_pct / 100.0,
        "Sale Price Mode": sale_mode,
        "Annual Property Appreciation %": appreciation_pct / 100.0,
        "Flat Sale Price (Override)": sale_price_override,
        "Closing Costs (Sale)": sale_closing_costs,
        "Sale Commission %": sale_commission_pct / 100.0,
        "Hold Period (years)": hold_years,
        "Number of Units": num_units,
        "Monthly Rent Total": monthly_rent,
        "Lead Name": st.session_state.get("lead_info", {}).get("name", ""),
        "Lead Email": st.session_state.get("lead_info", {}).get("email", ""),
        "Lead Phone": st.session_state.get("lead_info", {}).get("phone", "")
    }



    # Save calculation event as high-intent lead activity
    lead = st.session_state.get("lead_info", {})
    calc_event = {
        "purchase_price": purchase_price,
        "monthly_rent_total": monthly_rent,
        "hold_years": hold_years,
        "cap_rate": metrics.get("Cap Rate"),
        "cash_on_cash": metrics.get("Cash on Cash Return (Yr1)"),
        "dcr": metrics.get("DCR (Yr1)"),
        "total_profit_hold": metrics.get("Total Profit (hold)"),
        "number_of_units": num_units,
        "loan_term_years": loan_term_years,
        "use_buydown_2_1": use_buydown_2_1,
    }
    _save_feedback({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "event": "calculation_run",
        "name": lead.get("name", ""),
        "email": lead.get("email", ""),
        "phone": lead.get("phone", ""),
        **calc_event
    })
    _track_usage("calculation_run", calc_event)

    # --- Key Metrics (vertical) ---
    st.subheader("📊 Key Metrics & Predictions (Vertical)")
    display_rows = []
    red_metrics = set()

    dcr_val = metrics.get("DCR (Yr1)")
    if isinstance(dcr_val, (int, float)) and dcr_val < 1.0:
        red_metrics.add("DCR (Yr1)")
    be_val = metrics.get("Break-Even Occupancy % (Yr1)")
    if isinstance(be_val, (int, float)) and be_val > 1.0:
        red_metrics.add("Break-Even Occupancy % (Yr1)")

    for k, v in metrics.items():
        if isinstance(v, (int, float)) and v < 0:
            red_metrics.add(k)
        display_rows.append({"Metric": k, "Value": _fmt_metric(k, v)})
    display_df = pd.DataFrame(display_rows, columns=["Metric", "Value"])

    def _style_metrics_row(row):
        m = row["Metric"]
        val_style = "color:red;" if m in red_metrics else ""
        return ["", val_style]

    st.write(display_df.style.apply(_style_metrics_row, axis=1))

    # --- Year-by-Year Cashflows (formatted) ---
    st.subheader("📅 Yearly Cashflows (uses Annual Rental Appreciation %)")
    YCF_MONEY_COLS = [
        "Gross Rent (after vacancy) /mo", "Mgmt Fee /mo",
        "Other OpEx /mo (excl taxes/ins/HOA)", "Taxes /mo", "Insurance /mo", "HOA /mo",
        "Total OpEx /mo", "NOI (annual, prorated)",
        "Debt Service (annual, prorated)", "Cash Flow (annual)", "Cumulative Cash Flow"
    ]
    ycf_format = {c: "${:,.2f}" for c in YCF_MONEY_COLS if c in yearly_cf.columns}
    ycf_format.update({"Year": "{:.0f}", "Months Counted": "{:.0f}"})
    st.write(
        yearly_cf.style
            .format(ycf_format)
            .map(_style_red_neg, subset=[c for c in yearly_cf.columns if c in YCF_MONEY_COLS])
    )

    # --- DSCR Trend (table + line chart) ---
    st.subheader("📈 Debt Coverage Ratio (DCR) Trend by Year")
    if not yearly_cf.empty:
        dscr_df = yearly_cf[["Year", "NOI (annual, prorated)", "Debt Service (annual, prorated)"]].copy()
        dscr_df["DCR"] = dscr_df.apply(
            lambda r: (r["NOI (annual, prorated)"] / r["Debt Service (annual, prorated)"]) if r["Debt Service (annual, prorated)"] != 0 else np.nan,
            axis=1
        )
        st.write(
            dscr_df.style.format({
                "Year": "{:.0f}",
                "NOI (annual, prorated)": "${:,.2f}",
                "Debt Service (annual, prorated)": "${:,.2f}",
                "DCR": "{:.2f}"
            })
        )
        fig, ax = plt.subplots()
        ax.plot(dscr_df["Year"], dscr_df["DCR"], marker='o')
        ax.axhline(1.0, linestyle='--')
        ax.set_xlabel("Year")
        ax.set_ylabel("DCR")
        ax.set_title("Debt Coverage Ratio by Year")
        st.pyplot(fig)
    else:
        dscr_df = pd.DataFrame(columns=["Year", "NOI (annual, prorated)", "Debt Service (annual, prorated)", "DCR"])
        st.info("No yearly cashflow rows to chart.")

    # --- Amortization (formatted) ---
    st.subheader("📑 Amortization Schedule (Borrower vs Note) — first 200 rows")
    amort_view = amort.head(200)
    AMORT_MONEY_COLS = ["Note Payment", "Borrower Payment", "Subsidy", "Principal", "Interest", "Balance", "Equity"]
    amort_format = {c: "${:,.2f}" for c in AMORT_MONEY_COLS if c in amort_view.columns}
    amort_format.update({"Month": "{:.0f}"})
    st.write(
        amort_view.style
            .format(amort_format)
            .map(_style_red_neg, subset=AMORT_MONEY_COLS)
    )


    # --- Soft investor lead capture and investor tools (no upfront gate) ---
    _investor_deals_list_section(default_purchase_price=purchase_price)
    _email_roi_report_section(metrics, purchase_price, monthly_rent, num_units)
    _deal_review_section(metrics, purchase_price, monthly_rent, num_units)
    _investor_buy_box_section(metrics, purchase_price, monthly_rent, num_units)
    _save_compare_scenario_section(metrics, purchase_price, monthly_rent, num_units)
    _share_analysis_section(metrics, purchase_price, monthly_rent, num_units)
    _loan_quote_section(metrics, purchase_price, monthly_rent, num_units)
    _seller_lead_section(metrics, purchase_price, monthly_rent, num_units)
    _lead_score_section()

    # --- Feedback / follow-up request ---
    st.subheader("💬 Feedback & Deal Review")
    with st.form("feedback_form", clear_on_submit=True):
        usefulness = st.radio(
            "How useful was this calculator?",
            ["⭐⭐⭐⭐⭐ Excellent", "⭐⭐⭐⭐ Good", "⭐⭐⭐ Average", "⭐⭐ Needs Work"],
            horizontal=True
        )
        comments = st.text_area("What should I improve or add?")
        wants_review = st.checkbox("I want someone to review this deal with me")
        feedback_submit = st.form_submit_button("Submit Feedback", use_container_width=True)

    if feedback_submit:
        lead = st.session_state.get("lead_info", {})
        feedback_event = {
            "usefulness": usefulness,
            "comments": comments,
            "wants_review": wants_review,
            "purchase_price": purchase_price,
            "monthly_rent_total": monthly_rent,
            "cap_rate": metrics.get("Cap Rate"),
            "cash_on_cash": metrics.get("Cash on Cash Return (Yr1)"),
            "dcr": metrics.get("DCR (Yr1)")
        }
        _save_feedback({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "event": "feedback",
            "name": lead.get("name", ""),
            "email": lead.get("email", ""),
            "phone": lead.get("phone", ""),
            **feedback_event
        })
        _track_usage("feedback", feedback_event)
        st.success("Thank you! Your feedback was saved.")

    # --- Excel download ---
    excel_bytes = dataframes_to_xlsx_bytes(inputs, metrics, amort, yearly_cf, dscr_df)
    downloaded = st.download_button(
        "📥 Download Excel",
        data=excel_bytes,
        file_name="roi_amortization.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    if downloaded:
        lead = st.session_state.get("lead_info", {})
        download_event = {
            "purchase_price": purchase_price,
            "monthly_rent_total": monthly_rent,
            "cap_rate": metrics.get("Cap Rate"),
            "cash_on_cash": metrics.get("Cash on Cash Return (Yr1)"),
            "dcr": metrics.get("DCR (Yr1)")
        }
        _save_feedback({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "event": "excel_download",
            "name": lead.get("name", ""),
            "email": lead.get("email", ""),
            "phone": lead.get("phone", ""),
            **download_event
        })
        _track_usage("excel_download", download_event)

